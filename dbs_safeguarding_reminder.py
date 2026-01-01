#!/usr/bin/env python3

'''

#######################################
############ Description ##############
#######################################
This script integrates with the CHMeetings API to programmatically
retrieve member records and identify individuals whose DBS and Safeguarding
certifications have expired or are due to expire within one month.
It generates an Excel report containing the relevant expiry details for review
by the People Team Coordinator and Safeguarding Coordinator, and automatically
sends email notifications to affected members.

Deployment

The script can be deployed on GCP or AWS and executed on a scheduled basis
(for example, via Cloud Scheduler, EventBridge, or a cron job) to ensure
continuous monitoring and timely notifications.

'''

#######################################
############# Imports #################
#######################################

import requests                              # HTTP client for interacting with the CHMeetings REST API
from datetime import date, datetime, timedelta  # Date utilities for certificate expiry calculations
import logging                               # Structured logging for operational visibility and debugging
import os                                    # Access environment variables for configuration and secrets
import base64                                # Encode email content for Gmail API
from openpyxl import Workbook                # Create Excel workbooks in memory
from openpyxl.utils import get_column_letter # Utility for Excel column formatting
import io                                    # In-memory byte streams for Excel file generation
import sys                                   # System-level functions (e.g. controlled script exit)
from typing import List, Dict                # Type hints for improved readability and static analysis
from email.message import EmailMessage       # Construct MIME-compliant email messages
from google.oauth2.credentials import Credentials  # OAuth 2.0 credential handling for Google APIs
from googleapiclient.discovery import build  # Build Gmail API service client


# Define Script Name
CODENAME = "DBSSafeguardingReminder"

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logging.info(
    'Starting %s',
    CODENAME
)

# CHMeetings API Token
CHMEETINGS_API_TOKEN = os.getenv("CHMEETINGS_API_TOKEN")

# SG_FIELD_ID and DBS_FIELD_ID are defined by CHMeetings internally
# The values can be found in the raw API response of
# https://api.chmeetings.com/api/v1/people

DBS_FIELD_ID = os.getenv("DBS_FIELD_ID")

SG_FIELD_ID = os.getenv("SG_FIELD_ID")

# Google Client ID
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID")

# Google Client Secret
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET")

# Google Refresh Token
GOOGLE_REFRESH_TOKEN = os.getenv("GOOGLE_REFRESH_TOKEN")

# Google Sender
GMAIL_SENDER = os.getenv("GMAIL_SENDER")

# Google API Scope
SCOPES = ["https://www.googleapis.com/auth/gmail.send"]

# Email content
GMAIL_TO = os.getenv("GMAIL_TO", "")

EMAIL_SUBJECT = "Monthly DBS & Safeguarding Expiry Report"
EMAIL_BODY = """An Excel report is attached to this email. It contains the latest list of members whose DBS and/or Safeguarding certificates have expired or are due to expire within one month.

Members with expiring certificates have already received an email reminder with instructions on how to renew their certificates. You are advised to contact these members directly to assist them to proceed with the DBS and/or Safeguarding renewal process.
"""

EMAIL_BODY_HTML = """
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
    <p>An Excel report is attached to this email. It contains the latest list of members whose
    <strong>DBS</strong> and/or <strong>Safeguarding</strong> certificates have expired or are due to expire within one month.</p>
    <p>
        Members with expiring certificates have already received an email reminder with instructions
        on how to renew their certificates. You are advised to contact these members directly to assist
        them with the renewal process.
    </p>
</body>
</html>
"""


#######################################
############# Functions ###############
#######################################

def retrieve_members_info() -> tuple[list, list]:
    
    """Retrieve a list of members with information of first name, last name, email,
    Enhanced DBS and Safeguarding certificate expiry dates.
    """
    logging.info(
        "[retrieve_members_info] Retrieving members information"
    )
    
    # Initialise the members lists
    members_dbs = []
    members_sg = []
    
    # Set up variables for API Call - paginated response  
    session = requests.Session()
    session.headers.update({
        "apikey": CHMEETINGS_API_TOKEN,
        "accept": "application/json",
    })
    
    endpoint = "https://api.chmeetings.com/api/v1/people"  
    
    page = 1
    page_size = 25
    members = []
    
    # Loop through pages and extend the members list
    while True:
        params = {
            "page": page,
            "page_size": page_size
        }
        
        try:
            resp = session.get(
                endpoint,
                params=params,
                timeout=10)
        except requests.RequestException as e:
            logging.error(
                "[retrieving_members_info] HTTP request failed on page: %s: %s",
                page,
                e)
            sys.exit(1)
            
        if resp.status_code != 200:
            logging.error(
                "[retrieve_members_info] Unexpected status code %s on page %s. Response: %s",
                resp.status_code,
                page,
                resp.text[:500])
            sys.exit(1)
            
        # Now we are sure that API query is successful
        logging.info(
            "[retrieve_members_info] Successfully retrieve data on page %s from Chmeeting API",
            page
        )
        
        # Parse into json for further extraction
        data = resp.json()
        results = data.get("data", [])
        
        # Break the while loop when no more records are returned
        if not results:
            logging.info(
                "[retrieve_members_info] No more records found. Paginated API querying completes"
            )
            break
        
        # Add elements of results into members
        members.extend(results)
        
        logging.info(
            "[retrieve_members_info] Retrieved %d records from page %d",
            len(results),
            page)
        
        # Move on next page
        page += 1
        
        
    # Generate two lists of members with DBS and SG information respectively
    for member in members:
        
        # Place all key-value pairs of additional_field into a list
        additional_fields = member.get("additional_fields", [])
        
        # Initialised the matched flag
        matched_dbs = False
        matched_sg = False     
        safeguarding_expiry = None
        dbs_expiry = None
        
        # Loop through the additional fields of a member to check
        # the existence of DBS or Safeguarding expiry information
        for field in additional_fields:
            if str(field.get("field_id")) == SG_FIELD_ID:
                safeguarding_expiry = field.get("value")
                matched_sg = True
                
            if str(field.get("field_id")) == DBS_FIELD_ID:
                dbs_expiry = field.get("value")
                matched_dbs = True
                
                # Append the list members_dbs if DBS information is found
        if matched_dbs is True:
            member_dbs = {
                "id": member.get("id"),
                "first_name": member.get("first_name"),
                "last_name": member.get("last_name"),
                "email": member.get("email"),
                "mobile": member.get("mobile"),
                "dbs_expiry_date": dbs_expiry,
            }
            
            members_dbs.append(member_dbs)
            
        # Append the list members_sg if Safeguarding information is found
        if matched_sg is True:
            member_sg = {
                "id": member.get("id"),
                "first_name": member.get("first_name"),
                "last_name": member.get("last_name"),
                "email": member.get("email"),
                "mobile": member.get("mobile"),
                "safeguarding_expiry_date": safeguarding_expiry
            }
            
            members_sg.append(member_sg)
            
    return members_dbs, members_sg


def expiring_within_one_month(members: List[Dict], cert_type:str, days=30) -> list:
    
    """Filter the member list for DBS or safeguarding certificates that have expired
    or are due to expire within one month from the time the script is executed.
    Return the filtered list
    """
    
    logging.info(
        "[expiring_within_one_month] Filtering the members with expiring " + cert_type + " certs"
    )
    
    # Find out the cutoff date from today
    today = date.today()
    cutoff = today + timedelta(days=days)
    
    # Parse "YYYY-MM-DD" into a date object
    def parse_iso(d):
        try:
            return date.fromisoformat(d[:10]) if d else None
        except ValueError:
            return None
        
        # Initialise the expiring list
    expiring = []
    
    # Loop through the members and append those with expired or expiring DBS/Safeguarding certs
    for member in members:
        if cert_type == "dbs":
            dbs = parse_iso(member.get("dbs_expiry_date"))
            due = (dbs is not None) and (dbs <= cutoff)   # includes already expired
            
        if cert_type == "sg":
            sg  = parse_iso(member.get("safeguarding_expiry_date"))
            due = (sg is not None) and (sg <= cutoff)   # includes already expired
            
        if due:
            expiring.append(member)
            
    return expiring


def expiring_to_excel_bytes(rows: List[Dict], sheet_name: str = "Expiring Certificates") -> bytes:
    """
    Convert a list of dicts -> Excel (.xlsx) bytes in memory.
    """
    
    logging.info(
        "[expiring_to_excel_bytes] Converting the filtered list into Excel in memory"
    )
    
    if not rows:
        logging.info(
            "[expiring_to_csv_bytes] No list is found."
        )
        
        return b""
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    headers = list(rows[0].keys())
    ws.append(headers)
    
    for row in rows:
        ws.append([row.get(h) for h in headers])
        
        # Optional: auto-size columns
    for col_idx, header in enumerate(headers, start=1):
        max_len = max(len(str(header)), *(len(str(r.get(header, ""))) for r in rows))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
        
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def gmail_service_from_refresh_token():
    """Refreshing gmail service token"""
    
    logging.info(
        "[gmail_service_from_refresh_token] Refreshing gmail service token"
    )
    
    creds = Credentials(
        token=None,
        refresh_token=GOOGLE_REFRESH_TOKEN,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        scopes=SCOPES,
    )
    
    return build("gmail", "v1", credentials=creds, cache_discovery=False)


def send_email_with_excel(
    subject: str,
    body: str,
    to_addrs: list[str],
    attachments: list[tuple[str,bytes,str]],
    body_html: str | None = None
):
    
    """Sending gmail with Excel reports as attachment"""
    
    logging.info(
        "[send_email_with_excel] Sending gmail with excel reports as attachment"
    )
    
    sender = GMAIL_SENDER
    
    msg = EmailMessage()
    msg["To"] = ", ".join(to_addrs)
    msg["From"] = sender
    msg["Subject"] = subject
    
    # Plan text fallback (always)
    msg.set_content(body)
    
    # HTML version (optional)
    if body_html:
        msg.add_alternative(body_html, subtype="html")
        
        # Attachments
    for filename, file_bytes, mime_type in attachments:
        if not file_bytes:
            continue
        maintype, subtype = mime_type.split("/", 1)
        msg.add_attachment(file_bytes, maintype=maintype, subtype=subtype, filename=filename)
        
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    
    # Refresh the gmail API token
    service = gmail_service_from_refresh_token()
    
    # Send the email with verification of success
    try:
        response = service.users().messages().send(
            userId="me",
            body={"raw": raw}
        ).execute()
        
        message_id = response.get("id")
        
        if message_id:
            logging.info(
                "[send_email_with_excel] Email sent successfully. Message ID: %s",
                message_id,
            )
        else:
            logging.warning(
                "[send_email_with_excel] Email sent but no message ID returned"
            )
            
    except Exception as e:
        logging.exception(
            "[send_email_with_excel] Failed to send email"
        )
        raise
        
##############################################
############## MAIN APPLICATION ##############
##############################################
        
if __name__ == "__main__":
    logging.info("[main] Script running now...")
    
    # Retrieve members with DBS or Safeguarding cert
    member_info_dbs, member_info_sg = retrieve_members_info()
    
    # Filter out the members with expiring DBS certs
    expiring_member_dbs = expiring_within_one_month(member_info_dbs, "dbs", 30)
    
    # Filter out the members with expiring Safeguarding certs
    expiring_member_sg = expiring_within_one_month(member_info_sg, "sg", 30)
    
    # Store the expiring list as Excel in memory
    excel_bytes_dbs = expiring_to_excel_bytes(expiring_member_dbs)
    
    # Store the expiring list as Excel in memory
    excel_bytes_sg = expiring_to_excel_bytes(expiring_member_sg)
    
    # Place recipient address from the var which is delimited by comma
    to_addrs = [x.strip() for x in GMAIL_TO.split(",") if x.strip()]
    
    # Send the email with Excel MIME type
    XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    attachments = [
        ("DBS_Expiring_Report.xlsx", excel_bytes_dbs, XLSX_MIME),
        ("Safeguarding_Expiring_Report.xlsx", excel_bytes_sg, XLSX_MIME),
    ]
    
    send_email_with_excel(
        subject=EMAIL_SUBJECT,
        body=EMAIL_BODY,
        body_html=EMAIL_BODY_HTML,
        to_addrs=to_addrs,
        attachments=attachments,
    )
